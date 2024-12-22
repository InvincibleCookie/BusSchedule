import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment
from typing import List, Tuple, Dict, Any


def auto_adjust_column_width(sheet: Worksheet, padding: int = 2) -> None:
    """
    Автоматически подстраивает ширину столбцов листа Excel на основе содержимого ячеек.

    Args:
        sheet (Worksheet): Лист Excel, для которого необходимо подстроить ширину столбцов.
        padding (int, optional): Дополнительное пространство для ширины столбца. По умолчанию 2.
    """
    for col in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        max_length: int = 0
        col_letter: str = col[0].coordinate.split(":")[0][0]
        for cell in col:
            if cell.value:
                cell_length: int = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        adjusted_width: int = max_length + padding
        sheet.column_dimensions[col_letter].width = adjusted_width


def add_thick_border(cell: Any, border_side: str, thick_border: Border) -> None:
    """
    Добавляет толстую границу к указанной стороне ячейки.

    Args:
        cell (Cell): Ячейка Excel, к которой необходимо добавить границу.
        border_side (str): Сторона границы ('left', 'right', 'top', 'bottom').
        thick_border (Border): Объект Border с заданным стилем.
    """
    current_border: Border = cell.border
    border_kwargs = {'left': current_border.left, 'right': current_border.right, 'top': current_border.top,
                     'bottom': current_border.bottom, border_side: getattr(thick_border, border_side)}
    cell.border = Border(**border_kwargs)


def add_group_borders(sheet: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    """
    Добавляет толстые границы вокруг указанного диапазона ячеек.

    Args:
        sheet (Worksheet): Лист Excel, на котором добавляются границы.
        start_row (int): Начальный ряд диапазона.
        end_row (int): Конечный ряд диапазона.
        start_col (int): Начальный столбец диапазона.
        end_col (int): Конечный столбец диапазона.
    """
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            if row == start_row:
                cell.border = cell.border + Border(top=Side(style='thick'))
            if row == end_row:
                cell.border = cell.border + Border(bottom=Side(style='thick'))
            if col == start_col:
                cell.border = cell.border + Border(left=Side(style='thick'))
            if col == end_col:
                cell.border = cell.border + Border(right=Side(style='thick'))


def add_borders(sheet: Worksheet) -> None:
    """
    Добавляет только внешние толстые границы для групп ячеек на листе Excel.

    Args:
        sheet (Worksheet): Лист Excel, к которому добавляются границы.
    """
    thick_border: Border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
    )

    # Добавляем границы вокруг блока "Водитель" и "Вид смены"
    add_group_borders(sheet, start_row=1, end_row=2, start_col=1, end_col=2)

    # Определяем границы для дней недели
    days_start_col: int = 3
    total_days: int = 7

    for i in range(total_days):
        start_col: int = days_start_col + i * 3
        end_col: int = start_col + 2

        add_group_borders(sheet, start_row=1, end_row=sheet.max_row, start_col=start_col, end_col=end_col)


def add_summary_sheet(workbook: Workbook, job_result_df: pd.DataFrame, driver_columns: List[str]) -> None:
    """
    Добавляет лист 'Итоги' с агрегированной информацией о водителях.

    Args:
        workbook (Workbook): Книга Excel, в которую добавляется лист 'Итоги'.
        job_result_df (pd.DataFrame): DataFrame с результатами работы водителей.
        driver_columns (List[str]): Список столбцов, соответствующих водителям.
    """
    summary_sheet: Worksheet = workbook.create_sheet(title="Итоги", index=0)

    total_drivers: int = len(driver_columns)

    def contains_shift(data: Any, shift_time: str) -> bool:
        """
        Проверяет, содержит ли список или строка время смены.

        Args:
            data (Any): Данные ячейки (список или строка).
            shift_time (str): Время смены в формате 'HH:MM:SS'.

        Returns:
            bool: True, если время смены содержится, иначе False.
        """
        if isinstance(data, list):
            return any(f"Смена: {shift_time}" in str(item) for item in data)
        return False

    drivers_8_hour_shift: int = sum(
        job_result_df[col].apply(lambda x: contains_shift(x, "8:00:00")).any()
        for col in driver_columns
    )
    drivers_12_hour_shift: int = sum(
        job_result_df[col].apply(lambda x: contains_shift(x, "12:00:00")).any()
        for col in driver_columns
    )

    days_of_week: List[str] = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
    drivers_per_day: Dict[str, int] = {day: 0 for day in days_of_week}

    for col in driver_columns:
        for day_idx, day_name in enumerate(days_of_week):
            day_data = job_result_df[
                (job_result_df['Day'] == day_idx) & job_result_df[col].notna()
                ][col]
            if day_data.apply(lambda x: contains_shift(x, "8:00:00") or contains_shift(x, "12:00:00")).any():
                drivers_per_day[day_name] += 1
    summary_sheet.append(["Общие данные"])
    summary_sheet.append(["Общее количество водителей", total_drivers])
    summary_sheet.append(["Общее количество водителей с 8-часовой сменой", drivers_8_hour_shift])
    summary_sheet.append(["Общее количество водителей с 12-часовой сменой", drivers_12_hour_shift])
    summary_sheet.append([])

    summary_sheet.append(["Количество водителей по дням недели"])
    summary_sheet.append(["День недели", "Количество водителей"])
    for day, count in drivers_per_day.items():
        summary_sheet.append([day, count])

    auto_adjust_column_width(summary_sheet)


def excel_schedule(job_result_df: pd.DataFrame, output_file: str) -> None:
    """
    Создаёт Excel-файл с расписанием водителей и агрегированной информацией.

    Args:
        job_result_df (pd.DataFrame): DataFrame с результатами работы водителей.
        output_file (str): Путь к выходному Excel-файлу.
    """
    job_result_df = job_result_df.reset_index()
    job_result_df[['Day', 'Time']] = job_result_df['Time_index'].str.split(', ', expand=True)
    job_result_df['Day'] = job_result_df['Day'].astype(int)
    job_result_df['Time'] = job_result_df['Time'].str.strip()

    driver_columns: List[str] = [col for col in job_result_df.columns if
                                 col not in ['Time_index', 'placeholder', 'Day', 'Time']]

    days_of_week: List[str] = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

    workbook: Workbook = Workbook()
    add_summary_sheet(workbook, job_result_df, driver_columns)

    for driver_name in driver_columns:
        sheet: Worksheet = workbook.create_sheet(title=driver_name)

        sheet["A1"] = "Водитель"
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        sheet["B1"] = "Вид смены"
        sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        col_idx: int = 3
        for day in days_of_week:
            sheet.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 2)
            sheet.cell(row=1, column=col_idx).value = day
            sheet.cell(row=2, column=col_idx).value = "Автобус"
            sheet.cell(row=2, column=col_idx + 1).value = "Время"
            sheet.cell(row=2, column=col_idx + 2).value = "Действие"
            col_idx += 3

        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

        driver_data: pd.DataFrame = job_result_df[job_result_df[driver_name].notna()][['Day', 'Time', driver_name]]
        driver_data = driver_data.rename(columns={driver_name: 'Shift Details'})

        day_data: Dict[str, List[Tuple[str, str, str]]] = {day: [] for day in days_of_week}

        for _, row in driver_data.iterrows():
            current_time: str = row["Time"]
            shift_details: List[str] = row["Shift Details"] if isinstance(row["Shift Details"], list) else []
            action: str = shift_details[0] if len(shift_details) > 0 else ""
            bus: str = shift_details[2].split(": ")[-1] if len(shift_details) > 2 else ""

            day_idx: int = row["Day"]
            day_name: str = days_of_week[day_idx]

            day_data[day_name].append((bus, current_time, action))

        row_idx: int = 3
        sheet.cell(row=row_idx, column=1, value=driver_name)
        sheet.cell(row=row_idx, column=2, value=shift_details[1] if len(shift_details) > 1 else "")

        max_len: int = max(len(actions) for actions in day_data.values())

        for i in range(max_len):
            col_idx = 3
            for day in days_of_week:
                if i < len(day_data[day]):
                    bus, time, action = day_data[day][i]
                    sheet.cell(row=row_idx, column=col_idx, value=bus)
                    sheet.cell(row=row_idx, column=col_idx + 1, value=time)
                    sheet.cell(row=row_idx, column=col_idx + 2, value=action)
                col_idx += 3
            row_idx += 1

        auto_adjust_column_width(sheet)
        add_borders(sheet)

    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    workbook.save(output_file)
